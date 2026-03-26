#!/usr/bin/env python3
"""
gen_excel_report.py — 从 Markdown 月度报告生成 Excel 质量对比表

Usage:
    python3 gen_excel_report.py <reports_dir> [output_name]

Arguments:
    reports_dir   包含 *_community_quality_monthly_*.md 文件的目录（相对于项目根目录）
    output_name   可选，输出文件名（不含 .xlsx），默认取目录名

Example:
    python3 gen_excel_report.py reports/202603
    python3 gen_excel_report.py reports/202603 upstream_community_report_202603
"""

import sys
import os
import re
import json
import copy
from glob import glob
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", ".."))
TEMPLATE     = os.path.join(PROJECT_ROOT, "reports", "template", "report-template.xlsx")

# ── Display names ──────────────────────────────────────────────────────────────
DISPLAY_NAMES = {
    "pta": "PTA", "vllm": "vLLM", "triton": "Triton", "tilelang": "TileLang",
    "openeuler": "openEuler", "opengauss": "openGauss", "openubmc": "openUBMC",
    "mindspore": "MindSpore", "mindie": "MindIE", "mindstudio": "MindStudio",
    "mindspeed": "MindSpeed", "mindcluster": "MindCluster",
    "mindseriessdk": "MindSeriesSDK", "cann": "CANN", "cannopen": "CannOpen",
    "sgl": "SGL", "verl": "VeRL", "pytorch": "PyTorch", "ascendnpuir": "ascendnpuir",
    "boostkit": "BoostKit", "unifiedbus": "UnifiedBus", "openfuyao": "openFuyao",
}

# ── Colors ─────────────────────────────────────────────────────────────────────
C_TITLE  = "FF1E2761"
C_TEXT   = "FF1A1A1A"
C_BLUE   = "FF1A56AB"
C_GRAY   = "FF6B7280"
C_GREEN  = "FF009966"
C_RED    = "FFCC3333"
C_BG     = "FFFFFFFF"
C_BORDER = "FFBBBBBB"

# ── Style helpers ──────────────────────────────────────────────────────────────
def thin_border():
    s = Side(border_style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def std_fill():
    return PatternFill(fill_type="solid", fgColor=C_BG)

def std_align(h="center"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)

def set_cell(ws, row, col, value, bold=False, color=C_TEXT, halign="center", size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="等线", bold=bold, size=size, color=color)
    c.fill      = std_fill()
    c.alignment = std_align(halign)
    c.border    = thin_border()
    return c

# ── Parsing helpers ────────────────────────────────────────────────────────────
def extract_num(text):
    """Extract numeric value from markdown text like '299 人', '14.91 天', or '—'."""
    if text is None:
        return None
    text = text.strip()
    if text in ("—", "-", "", "N/A", "暂无数据", "无"):
        return None
    m = re.match(r"^([\d,]+\.?\d*)", text)
    if m:
        s = m.group(1).replace(",", "")
        try:
            return float(s) if "." in s else int(s)
        except ValueError:
            return None
    return None

def fmt_num(val):
    """Format a numeric value for display."""
    if val is None:
        return "—"
    if isinstance(val, float):
        return f"{val:,.2f}".rstrip("0").rstrip(".")
    if isinstance(val, int):
        return f"{val:,}"
    return str(val)

def build_cell_data(raw_curr, raw_prev, higher_is_better=True):
    """
    Returns (prev_str, curr_str, curr_bold, curr_color).
    higher_is_better=True  → increase is good (green), decrease is bad (red)
    higher_is_better=False → decrease is good (green), increase is bad (red)
    """
    curr_val = extract_num(raw_curr)
    prev_val = extract_num(raw_prev)

    if curr_val is None:
        return "—", "—", False, C_TEXT

    prev_str = fmt_num(prev_val)

    if prev_val is None:
        # No previous data — show current value only, no comparison
        return "—", fmt_num(curr_val), False, C_TEXT

    if prev_val == 0:
        pct = None
    else:
        pct = (curr_val - prev_val) / abs(prev_val) * 100

    curr_base = fmt_num(curr_val)

    if pct is None:
        return prev_str, curr_base, False, C_TEXT

    if abs(pct) < 0.05:
        curr_str = f"{curr_base} ▲+0.0%"
        is_good  = False   # flat = stagnation
    elif pct > 0:
        curr_str = f"{curr_base} ▲+{abs(pct):.1f}%"
        is_good  = higher_is_better
    else:
        curr_str = f"{curr_base} ▼-{abs(pct):.1f}%"
        is_good  = not higher_is_better

    return prev_str, curr_str, True, C_GREEN if is_good else C_RED

def parse_section(content, section_re):
    """Extract text of one section (until next ## or end)."""
    m = re.search(section_re, content, re.MULTILINE)
    if not m:
        return ""
    start = m.end()
    nxt = re.search(r"^##\s", content[start:], re.MULTILINE)
    return content[start: start + nxt.start()] if nxt else content[start:]

def parse_row(sec, label):
    """Extract (curr_raw, prev_raw) from a table row: | label | **curr** | prev | ... |"""
    pat = rf"\|\s*{re.escape(label)}\s*\|\s*\*\*(.+?)\*\*\s*\|\s*([^|]+?)\s*\|"
    m = re.search(pat, sec)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return "—", "—"

# ── Report parser ──────────────────────────────────────────────────────────────
def parse_report(filepath):
    """
    Parse a *_community_quality_monthly_*.md file.
    Returns dict:
        community   str   display name
        period      str   e.g. "202603"
        metrics     dict  key → (curr_raw, prev_raw)
    """
    with open(filepath, encoding="utf-8") as f:
        content = f.read()

    # Use only main report (before supplemental report)
    supp = re.search(r"^#+\s+.{0,30}补充报告", content, re.MULTILINE)
    if supp:
        content = content[: supp.start()]

    fname = os.path.basename(filepath)
    comm_key = re.match(r"^(.+?)_community_quality", fname)
    comm_key = comm_key.group(1) if comm_key else fname

    period_m = re.search(r"_(\d{6})\.md$", fname)
    period   = period_m.group(1) if period_m else ""

    display  = DISPLAY_NAMES.get(comm_key, comm_key)

    metrics = {}

    # 一、贡献活跃度
    s = parse_section(content, r"^##\s+一[、.]")
    metrics["activate"] = parse_row(s, "活跃开发者数")
    metrics["prs"]      = parse_row(s, "合入 PR 数")
    metrics["issues"]   = parse_row(s, "提交 Issue 数")

    # 二、代码审查质量
    s = parse_section(content, r"^##\s+二[、.]")
    metrics["review"]   = parse_row(s, "有效 Review 总数")

    # 三、适配集成引用度（特殊格式，无对比月）
    s = parse_section(content, r"^##\s+三[、.]")
    m = re.search(r"\|\s*适配 / 集成 / 引用项目总数\s*\|\s*\*\*(.+?)\*\*", s)
    metrics["integration"] = (m.group(1).strip() if m else "—", None)

    # 四、TOP 开发者留存
    s = parse_section(content, r"^##\s+四[、.]")
    metrics["top_dev"]  = parse_row(s, "TOP 开发者留存率")

    # 五、社区下载量
    s = parse_section(content, r"^##\s+五[、.]")
    metrics["download"] = parse_row(s, "年初至今下载量")

    # 六、Issue 响应效率（取平均值）
    s = parse_section(content, r"^##\s+六[、.]")
    metrics["issue_avg_first"] = parse_row(s, "平均首次响应时长")
    metrics["issue_avg_close"] = parse_row(s, "平均关闭时长")

    # 七、论坛响应效率
    s = parse_section(content, r"^##\s+七[、.]")
    metrics["forum_avg_first"] = parse_row(s, "平均首次响应时长")
    metrics["forum_avg_close"] = parse_row(s, "平均关闭时长")

    # 八、版本发布偏差
    s = parse_section(content, r"^##\s+八[、.]")
    metrics["version"]  = parse_row(s, "版本发布偏差")

    # 九、社区组织多样性
    s = parse_section(content, r"^##\s+九[、.]")
    metrics["companies"] = parse_row(s, "贡献组织数")

    # 十、主流平台搜索指数
    s = parse_section(content, r"^##\s+十[、.]")
    metrics["search"]   = parse_row(s, "平均搜索指数")

    return {
        "community": display,
        "community_key": comm_key,
        "period": period,
        "metrics": metrics,
    }

# ── Excel generation ───────────────────────────────────────────────────────────
# Row layout (row → (metric_key, higher_is_better, is_aggregate))
ROW_METRICS = {
    5:  None,                               # TTFHW — not available
    6:  None,                               # 严重缺陷 — not available
    7:  ("activate",        True,  False),
    8:  ("prs",             True,  False),
    9:  ("issues",          True,  False),
    10: ("review",          True,  False),
    11: ("integration",     True,  True),   # aggregate: no change calc
    12: ("top_dev",         True,  False),
    13: ("download",        True,  False),
    14: None,                               # CI 耗时 — not available
    15: None,                               # CI 利用率 — not available
    16: ("issue_avg_first", False, False),  # lower is better
    17: ("issue_avg_close", False, False),
    18: ("forum_avg_first", False, False),
    19: ("forum_avg_close", False, False),
    20: ("version",         False, False),
    21: ("companies",       True,  False),
    22: ("search",          True,  False),
}

# Col A/B/C label definitions
COL_A = {5: "社区生产力", 11: "社区创新力", 14: "社区稳健度"}
COL_B = {
    5: "软件质量", 7: "社区活跃度",
    11: "技术生态连接度", 12: "开发者留存度", 13: "下载量",
    14: "基础设施稳健度", 16: "治理完善度", 22: "社区技术影响力",
}
COL_C = {
    5:  ("典型场景TTFHW时间（了解/获取/使用/贡献）🔵", C_BLUE),
    6:  ("社区严重缺陷数",                             C_TEXT),
    7:  ("当期活跃开发者数",                           C_TEXT),
    8:  ("PR 数量",                                    C_TEXT),
    9:  ("Issue 数量",                                 C_TEXT),
    10: ("有效 Review 数量",                           C_TEXT),
    11: ("领域主流项目适配、集成、引用度",             C_TEXT),
    12: ("Top 开发者留存情况",                         C_TEXT),
    13: ("社区下载量（万次）",                         C_TEXT),
    14: ("CI 平均耗时（排队/准备/构建）（分钟）🔵",    C_BLUE),
    15: ("CI 资源利用率 🔵",                           C_BLUE),
    16: ("Issue 平均首次响应时间（天）",               C_TEXT),
    17: ("Issue 平均闭环时间（天）",                   C_TEXT),
    18: ("论坛问题平均首次响应时间（天）",             C_TEXT),
    19: ("论坛问题平均闭环时间（天）",                 C_TEXT),
    20: ("版本稳定发布偏差",                           C_TEXT),
    21: ("社区组织多样性",                             C_TEXT),
    22: ("主流平台搜索指数",                           C_TEXT),
}


def generate_excel(reports_dir, all_reports, output_path):
    """Generate the Excel file."""
    if not all_reports:
        raise ValueError("No community reports to process")

    n_comm     = len(all_reports)
    total_cols = 3 + n_comm * 2   # A-C + 2 cols per community

    # Infer stats period from first report
    # File is named YYYYMM where MM is the generation month;
    # the main report covers the *previous* natural month vs the month before that.
    period     = all_reports[0].get("period", "")
    year       = period[:4] if len(period) == 6 else ""
    gen_month  = int(period[4:6]) if len(period) == 6 else 0
    if gen_month > 1:
        stats_year    = year
        month_curr    = f"{gen_month - 1:02d}"
    else:
        stats_year    = str(int(year) - 1)
        month_curr    = "12"
    curr_month_int = int(month_curr)
    if curr_month_int > 1:
        month_prev = f"{curr_month_int - 1:02d}"
    else:
        month_prev = "12"
        stats_year = str(int(stats_year) - 1)

    today_str  = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Load template for style reference
    if not os.path.exists(TEMPLATE):
        raise FileNotFoundError(f"Template not found: {TEMPLATE}")
    tmpl = openpyxl.load_workbook(TEMPLATE)
    tws  = tmpl["综合指标对比"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "综合指标对比"

    # ── Row 1: Title ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value = "上游社区运营质量月度对比报告"
    src = tws["A1"]
    c.font = copy.copy(src.font); c.fill = copy.copy(src.fill)
    c.alignment = copy.copy(src.alignment)
    ws.row_dimensions[1].height = tws.row_dimensions[1].height

    # ── Row 2: Subtitle ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    c = ws["A2"]
    c.value = (f"统计周期：{stats_year}年{month_curr}月  |  "
               f"环比基准：{stats_year}年{month_prev}月  |  生成日期：{today_str}")
    src = tws["A2"]
    c.font = copy.copy(src.font); c.fill = copy.copy(src.fill)
    c.alignment = copy.copy(src.alignment)
    ws.row_dimensions[2].height = tws.row_dimensions[2].height

    # ── Row 3: Community name headers ───────────────────────────────────────────
    for col in range(1, 4):
        set_cell(ws, 3, col, None, bold=True, color=C_TITLE)
    ws.row_dimensions[3].height = tws.row_dimensions[3].height

    for i, report in enumerate(all_reports):
        sc = 4 + i * 2
        ws.merge_cells(f"{get_column_letter(sc)}3:{get_column_letter(sc+1)}3")
        src = tws["D3"]
        c = ws.cell(row=3, column=sc, value=report["community"])
        c.font      = copy.copy(src.font)
        c.fill      = copy.copy(src.fill)
        c.alignment = copy.copy(src.alignment)
        c.border    = thin_border()

    # ── Row 4: Column headers ───────────────────────────────────────────────────
    for col, lbl in enumerate(["指标分类", "牵引点", "原子指标"], start=1):
        set_cell(ws, 4, col, lbl, bold=True, color=C_TITLE)
    for i in range(n_comm):
        sc = 4 + i * 2
        set_cell(ws, 4, sc,   f"{month_prev}月", bold=True, color=C_TITLE)
        set_cell(ws, 4, sc+1, f"{month_curr}月", bold=True, color=C_TITLE)
    ws.row_dimensions[4].height = tws.row_dimensions[4].height

    # ── Rows 5-22: Labels + data ────────────────────────────────────────────────
    for row in range(5, 23):
        ws.row_dimensions[row].height = tws.row_dimensions[row].height

        # Col A
        set_cell(ws, row, 1, COL_A.get(row), bold=True, color=C_TEXT)

        # Col B
        set_cell(ws, row, 2, COL_B.get(row), bold=True, color=C_TEXT)

        # Col C
        c_lbl, c_color = COL_C[row]
        set_cell(ws, row, 3, c_lbl, bold=False, color=c_color, halign="left")

        # Data columns
        row_spec = ROW_METRICS[row]

        for i, report in enumerate(all_reports):
            sc = 4 + i * 2

            if row_spec is None:
                # Not tracked (TTFHW, CI, etc.)
                set_cell(ws, row, sc,   "—", color=C_GRAY)
                set_cell(ws, row, sc+1, "—", color=C_TEXT)
                continue

            key, higher, is_agg = row_spec
            curr_raw, prev_raw = report["metrics"].get(key, ("—", "—"))

            if is_agg:
                # Aggregate value — show same in both columns, no arrow
                val = fmt_num(extract_num(curr_raw)) if extract_num(curr_raw) is not None else "—"
                set_cell(ws, row, sc,   val, color=C_GRAY)
                set_cell(ws, row, sc+1, val, color=C_TEXT)
            else:
                prev_str, curr_str, curr_bold, curr_color = build_cell_data(
                    curr_raw, prev_raw, higher_is_better=higher
                )
                set_cell(ws, row, sc,   prev_str, color=C_GRAY)
                set_cell(ws, row, sc+1, curr_str, bold=curr_bold, color=curr_color)

    # ── Row 23: Footnote ────────────────────────────────────────────────────────
    ws.merge_cells(f"A23:{get_column_letter(total_cols)}23")
    c = ws["A23"]
    c.value = f"数据来源：datastat.osinfra.cn  |  报告由 Claude Code 自动生成  |  生成日期：{today_str}"
    src = tws["A23"]
    c.font = copy.copy(src.font); c.fill = copy.copy(src.fill)
    c.alignment = copy.copy(src.alignment)
    ws.row_dimensions[23].height = tws.row_dimensions[23].height

    # ── Column widths ────────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 15.3
    ws.column_dimensions["C"].width = 28
    for i in range(n_comm):
        sc = 4 + i * 2
        ws.column_dimensions[get_column_letter(sc)  ].width = 13
        ws.column_dimensions[get_column_letter(sc+1)].width = 18

    # ── Save ─────────────────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python3 gen_excel_report.py <reports_dir> [output_name]",
              file=sys.stderr)
        sys.exit(1)

    reports_dir = os.path.join(PROJECT_ROOT, sys.argv[1]) \
        if not os.path.isabs(sys.argv[1]) else sys.argv[1]

    if not os.path.isdir(reports_dir):
        print(f"ERROR: Directory not found: {reports_dir}", file=sys.stderr)
        sys.exit(1)

    # Auto-discover markdown reports
    pattern = os.path.join(reports_dir, "*_community_quality_monthly_*.md")
    md_files = sorted(glob(pattern))

    if not md_files:
        print(f"ERROR: No *_community_quality_monthly_*.md files found in {reports_dir}",
              file=sys.stderr)
        sys.exit(1)

    # Parse each report
    all_reports = []
    errors = []
    for fp in md_files:
        try:
            report = parse_report(fp)
            all_reports.append(report)
            print(f"  Parsed: {os.path.basename(fp)} → {report['community']}",
                  file=sys.stderr)
        except Exception as e:
            errors.append(f"{os.path.basename(fp)}: {e}")

    if errors:
        print("Warnings:", file=sys.stderr)
        for e in errors:
            print(f"  {e}", file=sys.stderr)

    if not all_reports:
        print("ERROR: No reports could be parsed", file=sys.stderr)
        sys.exit(1)

    # Determine output filename
    dir_name   = os.path.basename(reports_dir.rstrip("/"))
    output_name = sys.argv[2] if len(sys.argv) >= 3 else f"upstream_community_report_{dir_name}"
    output_path = os.path.join(reports_dir, output_name + ".xlsx")

    # Generate Excel
    generate_excel(reports_dir, all_reports, output_path)

    # Print JSON summary to stdout
    result = {
        "output":      os.path.relpath(output_path, PROJECT_ROOT),
        "communities": [r["community"] for r in all_reports],
        "period":      all_reports[0].get("period", ""),
        "errors":      errors,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
